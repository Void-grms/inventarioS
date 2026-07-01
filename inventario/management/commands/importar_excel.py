from datetime import datetime
from decimal import Decimal, InvalidOperation

from django.core.management.base import BaseCommand
from openpyxl import load_workbook

from inventario.models import Bien, Servicio

# Índices de columna (0-based) según el layout del Excel del inventario
COL = {
    "codigo": 0, "descripcion": 1, "responsable": 4, "valor": 10,
    "fecha_alta": 9, "ubicac": 13, "modelo": 18, "medidas": 20,
    "marca": 26, "estado": 28, "nro_serie": 33, "color": 38,
    "caracteristicas": 39,
}


def _txt(v):
    return "" if v is None else str(v).strip()


def _decimal(v):
    try:
        return Decimal(str(v))
    except (InvalidOperation, TypeError, ValueError):
        return None


def _fecha(v):
    if isinstance(v, datetime):
        return v.date()
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            return datetime.strptime(_txt(v), fmt).date()
        except ValueError:
            continue
    return None


class Command(BaseCommand):
    help = "Importa el inventario desde un archivo .xlsx"

    def add_arguments(self, parser):
        parser.add_argument("ruta", type=str)

    def handle(self, *args, **opts):
        wb = load_workbook(opts["ruta"], data_only=True)
        ws = wb.active
        creados, actualizados = 0, 0
        for fila in ws.iter_rows(min_row=2, values_only=True):
            codigo = _txt(fila[COL["codigo"]])
            if not codigo:
                continue
            nombre_serv = _txt(fila[COL["ubicac"]]) or "SIN UBICACION"
            servicio, _ = Servicio.objects.get_or_create(nombre=nombre_serv)
            estado = _txt(fila[COL["estado"]]) or "Regular"
            if estado not in dict(Bien.ESTADOS):
                estado = "Regular"
            defaults = dict(
                descripcion=_txt(fila[COL["descripcion"]]),
                servicio=servicio,
                estado=estado,
                responsable=_txt(fila[COL["responsable"]]),
                marca=_txt(fila[COL["marca"]]),
                modelo=_txt(fila[COL["modelo"]]),
                nro_serie=_txt(fila[COL["nro_serie"]]),
                color=_txt(fila[COL["color"]]),
                medidas=_txt(fila[COL["medidas"]]),
                caracteristicas=_txt(fila[COL["caracteristicas"]]),
                valor=_decimal(fila[COL["valor"]]),
                fecha_alta=_fecha(fila[COL["fecha_alta"]]),
            )
            _, creado = Bien.objects.update_or_create(
                codigo_patrimonial=codigo, defaults=defaults)
            creados += creado
            actualizados += not creado
        self.stdout.write(self.style.SUCCESS(
            f"Importados: {creados} nuevos, {actualizados} actualizados."))
