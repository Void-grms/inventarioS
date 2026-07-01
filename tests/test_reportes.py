import pytest
from openpyxl import load_workbook

from inventario.exports import bienes_por_servicio, bienes_por_responsable, exportar_excel


@pytest.fixture
def datos(db):
    from inventario.models import Bien, Servicio
    s1 = Servicio.objects.create(nombre="FARMACIA")
    s2 = Servicio.objects.create(nombre="TRIAJE")
    Bien.objects.create(codigo_patrimonial="A1", descripcion="x", servicio=s1,
                        estado="Bueno", responsable="ANA")
    Bien.objects.create(codigo_patrimonial="A2", descripcion="y", servicio=s1,
                        estado="Bueno", responsable="LUIS")
    Bien.objects.create(codigo_patrimonial="A3", descripcion="z", servicio=s2,
                        estado="Bueno", responsable="ANA")
    return None


def test_agrupacion_por_servicio(datos):
    grupos = bienes_por_servicio()
    nombres = {g["servicio"]: len(g["bienes"]) for g in grupos}
    assert nombres["FARMACIA"] == 2
    assert nombres["TRIAJE"] == 1


def test_agrupacion_por_responsable(datos):
    grupos = bienes_por_responsable()
    nombres = {g["responsable"]: len(g["bienes"]) for g in grupos}
    assert nombres["ANA"] == 2
    assert nombres["LUIS"] == 1


def test_exportar_excel_devuelve_bytes(datos, tmp_path):
    contenido = exportar_excel()
    ruta = tmp_path / "out.xlsx"
    ruta.write_bytes(contenido)
    wb = load_workbook(ruta)
    ws = wb.active
    assert ws.cell(row=1, column=1).value == "codigo_patrimonial"
    assert ws.max_row == 4  # encabezado + 3 bienes
